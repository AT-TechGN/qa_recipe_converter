import pytest
import io
from openpyxl import load_workbook
from unittest.mock import MagicMock
from apps.parser.excel_generator import ExcelGenerator


def make_mock_uc(order, uc_id, description, is_automated=False, status='À tester'):
    uc = MagicMock()
    uc.order = order
    uc.use_case_id = uc_id
    uc.description = description
    uc.preconditions = 'Prérequis'
    uc.steps = 'Étapes de test'
    uc.expected_results = 'Résultat attendu'
    uc.observed_results = ''
    uc.is_automated = is_automated
    uc.status = status
    return uc


def get_data_rows(ws, min_row=3):
    """Return rows with numeric first cell (actual UC data rows, not summary)."""
    return [
        row for row in ws.iter_rows(min_row=min_row)
        if row[0].value and isinstance(row[0].value, int)
    ]


class TestExcelGenerator:
    def test_generates_two_sheets(self):
        ucs = [make_mock_uc(1, 'UC001', 'Test login')]
        gen = ExcelGenerator(ucs, 'test.docx')
        buf = gen.generate()
        wb = load_workbook(buf)
        assert 'Use Cases' in wb.sheetnames
        assert 'Cas automatisé' in wb.sheetnames

    def test_use_cases_sheet_has_correct_count(self):
        ucs = [make_mock_uc(i, f'UC{i:03d}', f'Description {i}') for i in range(1, 6)]
        gen = ExcelGenerator(ucs, 'test.docx')
        buf = gen.generate()
        wb = load_workbook(buf)
        ws = wb['Use Cases']
        data_rows = get_data_rows(ws)
        assert len(data_rows) == 5

    def test_automated_sheet_filters_correctly(self):
        ucs = [
            make_mock_uc(1, 'UC001', 'Manuel', is_automated=False),
            make_mock_uc(2, 'UC002', 'Auto 1', is_automated=True),
            make_mock_uc(3, 'UC003', 'Auto 2', is_automated=True),
        ]
        gen = ExcelGenerator(ucs, 'test.docx')
        buf = gen.generate()
        wb = load_workbook(buf)
        ws = wb['Cas automatisé']
        data_rows = get_data_rows(ws)
        assert len(data_rows) == 2

    def test_empty_use_cases(self):
        gen = ExcelGenerator([], 'empty.docx')
        buf = gen.generate()
        wb = load_workbook(buf)
        assert 'Use Cases' in wb.sheetnames
        assert 'Cas automatisé' in wb.sheetnames

    def test_returns_bytes_io(self):
        ucs = [make_mock_uc(1, 'UC001', 'Test')]
        gen = ExcelGenerator(ucs, 'test.docx')
        buf = gen.generate()
        assert isinstance(buf, io.BytesIO)
        assert buf.tell() == 0  # seeked to start

    def test_header_row_content(self):
        ucs = [make_mock_uc(1, 'UC001', 'Test')]
        gen = ExcelGenerator(ucs, 'test.docx')
        buf = gen.generate()
        wb = load_workbook(buf)
        ws = wb['Use Cases']
        headers = [cell.value for cell in ws[2]]
        assert 'N°' in headers
        assert 'Use Case' in headers
        assert 'Description' in headers
        assert 'Résultats Attendus' in headers

    def test_source_filename_in_title(self):
        ucs = [make_mock_uc(1, 'UC001', 'Test')]
        gen = ExcelGenerator(ucs, 'ma_recette.docx')
        buf = gen.generate()
        wb = load_workbook(buf)
        ws = wb['Use Cases']
        title_cell = ws.cell(row=1, column=1).value
        assert 'ma_recette.docx' in (title_cell or '')

    def test_summary_row_exists(self):
        ucs = [make_mock_uc(i, f'UC{i:03d}', f'Desc') for i in range(1, 4)]
        gen = ExcelGenerator(ucs, 'test.docx')
        buf = gen.generate()
        wb = load_workbook(buf)
        ws = wb['Use Cases']
        # Find summary row (contains "Total")
        found = any(
            'Total' in str(cell.value)
            for row in ws.iter_rows()
            for cell in row
            if cell.value
        )
        assert found

    def test_no_automated_in_automated_sheet(self):
        ucs = [make_mock_uc(1, 'UC001', 'Manuel', is_automated=False)]
        gen = ExcelGenerator(ucs, 'test.docx')
        buf = gen.generate()
        wb = load_workbook(buf)
        ws = wb['Cas automatisé']
        data_rows = get_data_rows(ws)
        assert len(data_rows) == 0
