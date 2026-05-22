"""
Excel generator — QA Recipe Converter v2.1
Génère un fichier Excel professionnel avec :
  - En-tête personnalisée : nom entreprise, nom fichier, logo (droite)
  - Feuille "Use Cases" (tous les UC)
  - Feuille "Cas automatisé" (UC marqués is_automated=True)
"""
import io
import os
import base64
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

try:
    from PIL import Image as PILImage
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False

# ── Palette ───────────────────────────────────────────────────────────────────
COLOR_HEADER_BG   = "1E3A5F"   # Deep navy
COLOR_HEADER_FG   = "FFFFFF"   # White
COLOR_ALT_ROW     = "F0F4FA"   # Light blue-gray
COLOR_AUTOMATED   = "E8F5E9"   # Light green
COLOR_BORDER      = "B0BEC5"   # Light gray
COLOR_TITLE_BG    = "2D6A9F"   # Medium blue (feuille automatisée)
COLOR_COMPANY_BG  = "0F2342"   # Dark brand (header company row)
COLOR_COMPANY_FG  = "FFFFFF"
COLOR_META_BG     = "1A3A6E"   # Slightly lighter for second header row
COLOR_META_FG     = "C7D8F0"

HEADER_ROW_HEIGHT = 36
META_ROW_HEIGHT   = 22

COLUMNS = [
    ('use_case_id',      'Use Case',           18),
    ('description',      'Description',         40),
    ('preconditions',    'Préconditions',        35),
    ('steps',            'Étapes',              45),
    ('expected_results', 'Résultats Attendus',   40),
    ('observed_results', 'Résultats Observés',   40),
    ('status',           'Statut',              15),
]

NUM_COLS = len(COLUMNS) + 1   # +1 for N°


def _side(color=COLOR_BORDER):
    return Side(style='thin', color=color)


def _border():
    s = _side()
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type='solid')


class ExcelGenerator:
    def __init__(self, use_cases, source_filename: str = '',
                 company_name: str = '', excel_filename: str = '',
                 logo_path: str = None):
        self.use_cases       = list(use_cases)
        self.source_filename = source_filename
        self.company_name    = company_name or ''
        self.excel_filename  = excel_filename or source_filename
        self.logo_path       = logo_path      # absolute filesystem path or None
        self.wb              = Workbook()

    # ── Public API ─────────────────────────────────────────────────────────
    def generate(self) -> io.BytesIO:
        self._build_use_cases_sheet()
        self._build_automated_sheet()
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']
        buf = io.BytesIO()
        self.wb.save(buf)
        buf.seek(0)
        return buf

    # ── Sheet builders ─────────────────────────────────────────────────────
    def _build_use_cases_sheet(self):
        ws = self.wb.create_sheet("Use Cases")
        self._write_sheet(ws, self.use_cases, header_color=COLOR_HEADER_BG)

    def _build_automated_sheet(self):
        automated = [uc for uc in self.use_cases if uc.is_automated]
        ws = self.wb.create_sheet("Cas automatisé")
        self._write_sheet(ws, automated, header_color=COLOR_TITLE_BG)

    # ── Main writer ────────────────────────────────────────────────────────
    def _write_sheet(self, ws, use_cases, header_color: str):
        # ── Row 1 : company name ──────────────────────────────────────────
        company_text = self.company_name or 'QA Recipe Converter'
        ws.append([company_text])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NUM_COLS)
        c = ws.cell(row=1, column=1)
        c.fill      = _fill(COLOR_COMPANY_BG)
        c.font      = Font(bold=True, size=14, color=COLOR_COMPANY_FG, name='Calibri')
        c.alignment = Alignment(horizontal='left', vertical='center', indent=2)
        ws.row_dimensions[1].height = HEADER_ROW_HEIGHT

        # ── Row 2 : meta (source file + excel filename) ───────────────────
        meta_parts = []
        if self.source_filename:
            meta_parts.append(f"Source : {self.source_filename}")
        if self.excel_filename:
            meta_parts.append(f"Export : {self.excel_filename}")
        ws.append([" │ ".join(meta_parts)])
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NUM_COLS)
        c2 = ws.cell(row=2, column=1)
        c2.fill      = _fill(COLOR_META_BG)
        c2.font      = Font(size=9, color=COLOR_META_FG, italic=True, name='Calibri')
        c2.alignment = Alignment(horizontal='left', vertical='center', indent=2)
        ws.row_dimensions[2].height = META_ROW_HEIGHT

        # ── Logo (top-right of rows 1-2) ──────────────────────────────────
        self._insert_logo(ws, target_row=1, target_col=NUM_COLS)

        # ── Row 3 : column headers ────────────────────────────────────────
        ws.append(['N°'] + [col[1] for col in COLUMNS])
        h_fill  = _fill(header_color)
        h_font  = Font(bold=True, color=COLOR_HEADER_FG, size=10, name='Calibri')
        border  = _border()
        for cell in ws[3]:
            cell.fill      = h_fill
            cell.font      = h_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border    = border
        ws.row_dimensions[3].height = 30

        # Column widths
        ws.column_dimensions['A'].width = 6
        for i, (_, _, w) in enumerate(COLUMNS):
            ws.column_dimensions[get_column_letter(i + 2)].width = w

        # Freeze below header
        ws.freeze_panes = 'A4'

        # ── Data rows (start at row 4) ────────────────────────────────────
        alt_fill  = _fill(COLOR_ALT_ROW)
        auto_fill = _fill(COLOR_AUTOMATED)

        for row_num, uc in enumerate(use_cases, start=1):
            row_data = [row_num] + [
                uc.status if field == 'status' else getattr(uc, field, '')
                for field, _, _ in COLUMNS
            ]
            ws.append(row_data)
            excel_row = row_num + 3   # title(1) + meta(1) + header(1) = offset 3

            if uc.is_automated:
                fill = auto_fill
            elif row_num % 2 == 0:
                fill = alt_fill
            else:
                fill = None

            for col_idx in range(1, len(row_data) + 1):
                cell            = ws.cell(row=excel_row, column=col_idx)
                cell.border     = border
                cell.alignment  = Alignment(
                    vertical='top', wrap_text=True,
                    horizontal='center' if col_idx == 1 else 'left'
                )
                cell.font = Font(bold=(col_idx == 1), size=10, name='Calibri')
                if fill:
                    cell.fill = fill
            ws.row_dimensions[excel_row].height = 60

        # Auto-filter
        if use_cases:
            last_col = get_column_letter(NUM_COLS)
            ws.auto_filter.ref = f"A3:{last_col}{len(use_cases) + 3}"

        # Summary
        if use_cases:
            summary_row = len(use_cases) + 5
            sc = ws.cell(row=summary_row, column=1,
                         value=f"Total : {len(use_cases)} cas de tests")
            sc.font = Font(bold=True, italic=True, color="666666", size=9)

    # ── Logo helper ────────────────────────────────────────────────────────
    def _insert_logo(self, ws, target_row: int, target_col: int):
        """Insert logo image anchored to top-right of header area."""
        if not self.logo_path:
            return
        if not os.path.exists(self.logo_path):
            return

        ext = os.path.splitext(self.logo_path)[1].lower()
        # SVG not supported by openpyxl directly; skip gracefully
        if ext == '.svg':
            return

        try:
            # Resize logo to fit 2 header rows height (≈ 58 px)
            logo_path_to_use = self.logo_path

            if PIL_AVAILABLE:
                img_pil = PILImage.open(self.logo_path)
                # Convert palette/RGBA to RGB for JPEG compatibility
                if img_pil.mode in ('RGBA', 'P', 'LA'):
                    bg = PILImage.new('RGB', img_pil.size, (255, 255, 255))
                    if img_pil.mode == 'P':
                        img_pil = img_pil.convert('RGBA')
                    bg.paste(img_pil, mask=img_pil.split()[-1] if img_pil.mode == 'RGBA' else None)
                    img_pil = bg

                # Resize keeping aspect ratio
                target_h = 52  # pixels
                ratio     = target_h / img_pil.height
                new_w     = int(img_pil.width * ratio)
                img_pil   = img_pil.resize((new_w, target_h), PILImage.LANCZOS)

                buf_img = io.BytesIO()
                img_pil.save(buf_img, format='PNG')
                buf_img.seek(0)
                logo_path_to_use = buf_img

            xl_img = XLImage(logo_path_to_use)
            # Anchor: column = target_col letter, row = 1
            col_letter = get_column_letter(target_col)
            xl_img.anchor = f"{col_letter}1"
            ws.add_image(xl_img)

        except Exception as e:
            # Never crash the Excel generation because of a logo issue
            import logging
            logging.getLogger(__name__).warning(f"Logo insert failed: {e}")
